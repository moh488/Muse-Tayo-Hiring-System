import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def export_candidates_to_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Muuse Tayo Candidates'

    brand_blue_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')

    headers = ['Ref Code', 'Candidate Name', 'Applied Position', 'Department', 'Current Stage', 'Application Date']
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = brand_blue_fill
        cell.alignment = center_align

    for item in queryset:
        row_data = [
            item.job.reference_code,
            f"{item.candidate.first_name} {item.candidate.last_name}",
            item.job.title,
            item.job.get_department_display(),
            item.get_stage_display(),
            item.applied_date.strftime('%Y-%m-%d'),
        ]
        ws.append(row_data)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Muuse_Tayo_Hiring_Export.xlsx"'
    wb.save(response)
    return response


def export_jobs_to_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Job Requisitions'
    headers = ['Reference', 'Title', 'Department', 'Status', 'Location', 'Deadline', 'Created']
    ws.append(headers)
    for job in queryset:
        ws.append([
            job.reference_code, job.title, job.get_department_display(),
            job.get_status_display(), job.location,
            job.deadline.strftime('%Y-%m-%d'), job.created_at.strftime('%Y-%m-%d'),
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Muuse_Tayo_Jobs_Export.xlsx"'
    wb.save(response)
    return response


def export_summary_pdf(title, headers, rows):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    if rows:
        data = [headers] + rows
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ]))
        elements.append(table)
    doc.build(elements)
    return response
